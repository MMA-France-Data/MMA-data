"""
exporter_coachs.py — LE TABLEAU DES DIALOGUES DE COACH, POUR MAEL.

    node ../tableau/vider_coachs.js > coach_rows.json  (fait par lancer.sh)
    python3 exporter_coachs.py coach_rows.json dialogues_coachs.xlsx

Mael (03/09) : « tu me sors un tableau avec toutes les repliques et
reponses, et les consequences, je te les modifie et je te renvoie ».
Une ligne par REPONSE. Les cases jaunes sont a lui, les grises portent
la cle (ID) qui permet de remettre chaque ligne a sa place au retour.
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

src, dst = sys.argv[1], sys.argv[2]
J = json.load(open(src, encoding="utf-8"))
rows = J["rows"]
SUJ = {s["cle"]: s["lab"] for s in J["SUJETS"]}
MOM = {"bureau": "Le bureau (tu vas le voir)", "bord_du_tapis": "Au bord du tapis",
       "debrief": "Le débrief du lendemain", "accrochage": "L'accrochage", "porte": "La porte (il pense à partir)"}
# /!\ CES DEUX TABLES SONT LA TRADUCTION ALLER ; l'importateur les inverse.
from exporter_coachs_tables import EFF, DECL
VIE = {"courante": "souvent", "saison": "1×/an", "unique": "1×"}

wb = Workbook()
ws = wb.active; ws.title = "Dialogues"
F = Font(name="Arial", size=10); FB = Font(name="Arial", size=10, bold=True, color="FFFFFF")
hdr = PatternFill("solid", fgColor="1F3A5F"); gris = PatternFill("solid", fgColor="EDEDED"); jaune = PatternFill("solid", fgColor="FFF6CC")
bord = Side(style="thin", color="BBBBBB")
cols = [("ID", 9), ("Moment", 18), ("Sujet", 26), ("Quand ça sort", 28), ("Pour qui (voix)", 22), ("Revient", 9),
        ("CE QU'IL DIT", 60), ("TA RÉPONSE", 34), ("SA RÉACTION", 60), ("Entente", 8), ("Ton", 11), ("CONSÉQUENCE", 36)]
for i, (n, w) in enumerate(cols, 1):
    c = ws.cell(row=1, column=i, value=n); c.font = FB; c.fill = hdr
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30
prev = None
for k, r in enumerate(rows, 2):
    scene = r["id"].split("#")[0]
    vals = [r["id"], MOM[r["moment"]], SUJ.get(r["sujet"], ""), DECL.get(r["si"], r["si"]),
            r["voix"] or "tout le monde", VIE[r["vie"]],
            r["texte"] if scene != prev else "", r["lab"], r["r"], r["d"], r["ton"], EFF.get(r["effet"], "")]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=k, column=i, value=v); c.font = F
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = Border(bottom=bord)
        c.fill = gris if i <= 6 else jaune
    prev = scene
dv = DataValidation(type="whole", operator="between", formula1="-6", formula2="6", showErrorMessage=True,
                    errorTitle="Entente", error="Entre -6 et +6 : une phrase ne fait pas basculer une relation.")
ws.add_data_validation(dv); dv.add(f"J2:J{len(rows) + 1}")
ws2 = wb.create_sheet("Listes")
ws2["A1"] = "CONSÉQUENCES possibles"; ws2["A1"].font = FB; ws2["A1"].fill = hdr
for i, v in enumerate(EFF.values(), 2): ws2.cell(row=i, column=1, value=v).font = F
ws2.column_dimensions["A"].width = 48
dv2 = DataValidation(type="list", formula1=f"=Listes!$A$2:$A${len(EFF) + 1}", allow_blank=True, showErrorMessage=True,
                     errorTitle="Conséquence", error="Choisis dans la liste — une conséquence inventée ne ferait rien.")
ws.add_data_validation(dv2); dv2.add(f"L2:L{len(rows) + 1}")

ws0 = wb.create_sheet("LIS-MOI", 0); ws0.column_dimensions["A"].width = 100
notes = [
    ("CE QUE TU PEUX MODIFIER (les cases jaunes)", True),
    ("• CE QU'IL DIT — la phrase du coach. Elle n'est écrite que sur la première ligne de la scène ; les lignes suivantes sont les autres réponses de la même scène.", False),
    ("• TA RÉPONSE — le bouton que tu cliques.", False),
    ("• SA RÉACTION — ce qu'il répond.", False),
    ("• Entente — de -6 à +6. Ce que la réponse fait à votre relation. Négatif = ça coûte.", False),
    ("• Ton — un mot libre (calme, sec, franc, dur, chaud…). Sert au journal.", False),
    ("• CONSÉQUENCE — choisis dans la liste déroulante. Vide = la réponse ne fait rien d'autre que bouger l'entente.", False),
    ("", False),
    ("CE QUE TU NE TOUCHES PAS (les cases grises)", True),
    ("• ID — la clé qui me permet de remettre ta ligne au bon endroit. Si tu la changes, je perds la ligne.", False),
    ("• Moment, Sujet, Quand ça sort, Pour qui, Revient — dis-moi à côté si tu veux les changer, je le ferai à la main.", False),
    ("", False),
    ("LES RÈGLES QUE LE JEU VÉRIFIE (sinon la ligne est refusée, et je te dis laquelle)", True),
    ("• Aucun chiffre dans les textes (écris « trois », pas « 3 »). Un coach ne parle pas en pourcentages.", False),
    ("• Une réponse qui donne de l'argent doit le DIRE dans TA RÉPONSE (salaire, augmentation, tarif…).", False),
    ("• {gars} = le nom du combattant posé sur la table. {autre} = quelqu'un d'autre de la salle. Garde-les tels quels, le jeu les remplit.", False),
    ("• Une réaction peut être courte (« Il est vexé. » passe) ; la phrase du coach fait au moins 40 caractères.", False),
    ("", False),
    ("POUR AJOUTER une réponse : insère une ligne, mets l'ID de la scène suivi de #5 (ou #6). Pour SUPPRIMER : efface TA RÉPONSE, je la retire.", False),
    ("Renvoie-moi le fichier tel quel : je le rentre dans le jeu et je te dis ce qui a été refusé et pourquoi.", False),
]
for i, (t, b) in enumerate(notes, 1):
    c = ws0.cell(row=i, column=1, value=t); c.font = Font(name="Arial", size=11, bold=b); c.alignment = Alignment(wrap_text=True)
wb.save(dst)
print(f"{dst} : {len(rows)} lignes")
