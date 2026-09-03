"""
importer_coachs.py — LE RETOUR DU TABLEAU DE MAEL.

    python3 importer_coachs.py dialogues_coachs.xlsx retours.json

Lit les cases jaunes, retraduit les libelles en cles du jeu, et ecrit un
JSON que remettre_coachs.js applique. Ne juge rien : les regles (chiffres,
longueurs, argent annonce) sont celles du banc 39, qui tranche apres.
"""
import json, sys
from openpyxl import load_workbook
from exporter_coachs_tables import EFF, DECL  # les memes tables, inversees ici

src, dst = sys.argv[1], sys.argv[2]
EFF_INV = {v.strip().lower(): k for k, v in EFF.items()}
wb = load_workbook(src, data_only=True)
ws = wb["Dialogues"]
hdr = [c.value for c in ws[1]]
col = {n: i for i, n in enumerate(hdr)}
out, refus = [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[col["ID"]]: continue
    rid = str(row[col["ID"]]).strip()
    if "#" not in rid: refus.append(f"{rid} : ID sans #"); continue
    lab = row[col["TA RÉPONSE"]]
    eff_txt = (row[col["CONSÉQUENCE"]] or "")
    eff = EFF_INV.get(str(eff_txt).strip().lower(), None) if str(eff_txt).strip() else ""
    if eff is None: refus.append(f"{rid} : conséquence inconnue « {eff_txt} »"); eff = ""
    d = row[col["Entente"]]
    try: d = int(d) if d not in (None, "") else 0
    except Exception: refus.append(f"{rid} : entente « {d} » n'est pas un nombre"); d = 0
    out.append({"id": rid, "texte": (row[col["CE QU'IL DIT"]] or "") or None,
                "lab": (lab or "").strip() if lab else "", "r": (row[col["SA RÉACTION"]] or "").strip(),
                "d": d, "ton": (row[col["Ton"]] or "neutre").strip(), "effet": eff,
                "supprimer": not (lab and str(lab).strip())})
json.dump({"lignes": out, "refus": refus}, open(dst, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"{len(out)} lignes lues · {len(refus)} refusées")
for r in refus[:10]: print("  ✗", r)
